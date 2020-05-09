import openpyxl, os, sys

#os.chdir('C:\\Users\\Szymon\\Documents\\Python file\\pythonAttemptTwo\\arkusze')

wb = openpyxl.Workbook()
sheet = wb.active

if len(sys.argv) < 2:
    num = 6
else:
    num = int(sys.argv[1])

fontObj = openpyxl.styles.Font(bold=True)

for rw in range(1, num+1):
    sheet['A' + str(rw+1)] = rw
    sheet['A' + str(rw+1)].font = fontObj
    sheet[openpyxl.utils.cell.get_column_letter(rw+1) + '1'] = rw
    sheet[openpyxl.utils.cell.get_column_letter(rw+1) + '1'].font= fontObj
    for col in range(1, num+1):
        sheet.cell(row=rw+1, column=col+1).value = rw*col


wb.save('multiplicationTable.xlsx')
