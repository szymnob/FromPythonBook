import openpyxl, os, sys

def blankInsert(fileIn, fileOut, spaceStart, spaceHeight):
    old_wb = openpyxl.load_workbook(fileIn)
    old_sheet = old_wb.active

    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    max_column = old_sheet.max_column
    max_row = old_sheet.max_row

    for col in range(1, max_column+1):
        for height in range(1, max_row+1):
            if height < spaceStart:
                new_sheet.cell(column=col, row=height).value = old_sheet.cell(column=col, row=height).value
            else:
                new_sheet.cell(column=col, row=height+spaceHeight).value = old_sheet.cell(column=col, row=height).value

    new_wb.save(fileOut)

#for test purposes
os.chdir('C:\\Users\\Szymon\\Documents\\Python file\\pythonAttemptTwo\\arkusze')

if len(sys.argv) < 4:
    n = 4
    m = 3
    In = 'test.xlsx'
    Out = 'blankRow.xlsx'
else:
    n = int(sys.argv[1])
    m = int(sys.argv[2])
    In = sys.argv[3]
    Out = sys.argv[4]

blankInsert(In, Out, n, m)