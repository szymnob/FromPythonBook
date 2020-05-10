import openpyxl, os, sys

os.chdir('C:\\Users\\Szymon\\Documents\\Python file\\pythonAttemptTwo\\arkusze')

old_wb = openpyxl.load_workbook('test.xlsx')
old_sheet = old_wb.active

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

max_column = old_sheet.max_column
max_row = old_sheet.max_row

mainList = []
helpList = []

for col in range(1, max_column+1):
    for height in range(1, max_row+1):
        helpList.append(old_sheet.cell(column=col, row=height).value)    
    mainList.append(helpList)
    helpList = []

for col in range(len(mainList)):
    for rw in range(len(mainList[col])):
        new_sheet.cell(row=col+1, column=rw+1).value = mainList[col][rw]


new_wb.save('amovedCells.xlsx')


