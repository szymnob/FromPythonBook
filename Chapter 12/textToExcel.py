import openpyxl, os, sys

os.chdir('C:\\Users\\Szymon\\Documents\\Python file\\pythonAttemptTwo\\arkusze')

wb = openpyxl.Workbook()
sheet = wb.active

files = ['1.txt', '2.txt', '3.txt']

for col in range(len(files)):
    sheet.cell(column=col+1, row=1).value = files[col]
    lines = open(files[col], 'r').readlines()
    for rw in range(len(lines)):
        sheet.cell(column=col+1, row=rw+2).value = lines[rw].strip('\n')


wb.save('atext.xlsx')
